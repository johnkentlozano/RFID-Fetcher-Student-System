import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import matplotlib.pyplot as plt
from collections import Counter
from utils.database import db_connect


class TeacherDownloadFrame(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg="#F0F4F8")

        self.controller = controller
        self.all_rows = []

        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview", rowheight=28, font=("Segoe UI", 10), background="white", fieldbackground="white")
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#0047AB", foreground="white")
        style.map("Treeview", background=[("selected", "#0047AB")], foreground=[("selected", "white")])

        # ================= HEADER =================
        header = tk.Frame(self, bg="#0047AB", height=80)
        header.pack(fill="x")
        header.pack_propagate(False)

        tk.Label(
            header,
            text="📥 DOWNLOAD STUDENT LOGS",
            font=("Helvetica", 20, "bold"),
            bg="#0047AB",
            fg="white"
        ).pack(side="left", padx=20, pady=15)

        self.status_label = tk.Label(
            header,
            text="",
            font=("Helvetica", 10),
            bg="#0047AB",
            fg="#B0C4DE"
        )
        self.status_label.pack(side="right", padx=20)

        # ================= MAIN BODY =================
        main_body = tk.Frame(self, bg="#F0F4F8")
        main_body.pack(fill="both", expand=True, padx=20, pady=10)

        # ================= LEFT: FILTER PANEL =================
        left_col = tk.Frame(main_body, bg="#F0F4F8", width=260)
        left_col.pack(side="left", fill="y", padx=(0, 15))
        left_col.pack_propagate(False)

        # Filter card
        filter_card = tk.Frame(left_col, bg="white", highlightthickness=1, highlightbackground="#D1D9E6")
        filter_card.pack(fill="x", pady=(0, 12), ipady=5)

        tk.Label(filter_card, text="FILTER OPTIONS", font=("Arial", 10, "bold"),
                 bg="white", fg="#0047AB").grid(row=0, column=0, columnspan=2, padx=15, pady=(12, 8), sticky="w")

        tk.Label(filter_card, text="Start Date:", font=("Arial", 9, "bold"),
                 bg="white", fg="#374151").grid(row=1, column=0, padx=15, pady=5, sticky="w")
        self.start_date = DateEntry(filter_card, date_pattern='yyyy-mm-dd',
                                    font=("Arial", 10), background="#0047AB", foreground="white")
        self.start_date.grid(row=1, column=1, padx=10, pady=5, sticky="ew")

        tk.Label(filter_card, text="End Date:", font=("Arial", 9, "bold"),
                 bg="white", fg="#374151").grid(row=2, column=0, padx=15, pady=5, sticky="w")
        self.end_date = DateEntry(filter_card, date_pattern='yyyy-mm-dd',
                                  font=("Arial", 10), background="#0047AB", foreground="white")
        self.end_date.grid(row=2, column=1, padx=10, pady=5, sticky="ew")

        tk.Label(filter_card, text="Student:", font=("Arial", 9, "bold"),
                 bg="white", fg="#374151").grid(row=3, column=0, padx=15, pady=5, sticky="w")
        self.student_combo = ttk.Combobox(filter_card, state="readonly", font=("Arial", 10))
        self.student_combo.grid(row=3, column=1, padx=10, pady=5, sticky="ew")

        tk.Label(filter_card, text="Live Search:", font=("Arial", 9, "bold"),
                 bg="white", fg="#374151").grid(row=4, column=0, padx=15, pady=5, sticky="w")
        self.search_var = tk.StringVar()
        tk.Entry(filter_card, textvariable=self.search_var,
                 font=("Arial", 10), relief="solid", bd=1).grid(row=4, column=1, padx=10, pady=5, sticky="ew")
        self.search_var.trace_add("write", self.live_search)

        filter_card.columnconfigure(1, weight=1)

        # ================= ACTION BUTTONS =================
        btn_card = tk.Frame(left_col, bg="white", highlightthickness=1, highlightbackground="#D1D9E6")
        btn_card.pack(fill="x", pady=(0, 12), ipady=8)

        tk.Label(btn_card, text="ACTIONS", font=("Arial", 10, "bold"),
                 bg="white", fg="#0047AB").pack(padx=15, pady=(12, 8), anchor="w")

        buttons = [
            ("🔍  Load Students", "#0047AB", self.load_students),
            ("📊  Load Data",     "#2196F3", self.load_data),
            ("📁  Export Excel",  "#4CAF50", self.export_excel),
            ("📈  Show Chart",    "#F59E0B", self.show_chart),
        ]
        for text, color, cmd in buttons:
            tk.Button(
                btn_card, text=text, bg=color, fg="white",
                font=("Arial", 9, "bold"), width=22, bd=0,
                pady=6, cursor="hand2", command=cmd
            ).pack(padx=15, pady=3)

        # ================= STATS CARD =================
        self.stats_card = tk.Frame(left_col, bg="white", highlightthickness=1, highlightbackground="#D1D9E6")
        self.stats_card.pack(fill="x", ipady=8)

        tk.Label(self.stats_card, text="📌 SUMMARY", font=("Arial", 10, "bold"),
                 bg="white", fg="#0047AB").pack(padx=15, pady=(12, 6), anchor="w")

        self.total_label   = tk.Label(self.stats_card, text="Total Records: —", font=("Arial", 9), bg="white", fg="#374151")
        self.total_label.pack(padx=15, anchor="w")
        self.student_count = tk.Label(self.stats_card, text="Unique Students: —", font=("Arial", 9), bg="white", fg="#374151")
        self.student_count.pack(padx=15, anchor="w", pady=(2, 8))

        # ================= RIGHT: TABLE =================
        right_col = tk.Frame(main_body, bg="#F0F4F8")
        right_col.pack(side="right", fill="both", expand=True)

        tk.Label(right_col, text="Fetch History", font=("Arial", 12, "bold"),
                 bg="#F0F4F8", fg="#0047AB").pack(anchor="w", pady=(0, 5))

        table_frame = tk.Frame(right_col, bg="white", highlightthickness=1, highlightbackground="#D1D9E6")
        table_frame.pack(fill="both", expand=True)

        columns = ("Student Name", "Student ID", "Time Out", "Fetcher", "Location")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings")

        col_widths = {"Student Name": 180, "Student ID": 100, "Time Out": 150, "Fetcher": 150, "Location": 120}
        for col in columns:
            self.tree.heading(col, text=col.upper())
            self.tree.column(col, anchor="center", width=col_widths.get(col, 130))

        # Alternating row colors
        self.tree.tag_configure("odd",  background="#F8FAFC")
        self.tree.tag_configure("even", background="white")

        vsb = ttk.Scrollbar(table_frame, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        # ================= AUTO LOAD =================
        self.refresh_all()

    # ================= AUTO REFRESH =================
    def refresh_all(self):
        self.load_students()
        self.load_data()
        self.after(3000, self.refresh_all)

    # ================= LOAD STUDENTS =================
    def load_students(self):
        # FIX: guard against None current_user
        if not self.controller.current_user:
            return
        try:
            employee_id = self.controller.current_user.get("employee_id")
            if not employee_id:
                return

            with db_connect() as conn:
                with conn.cursor() as cur:
                    cur.execute("""
                        SELECT s.Student_name
                        FROM classroom c
                        JOIN student s ON c.student_id = s.Student_id
                        WHERE c.employee_id = %s
                        ORDER BY s.Student_name
                    """, (employee_id,))
                    students = [row[0] for row in cur.fetchall()]

            self.student_combo["values"] = ["ALL"] + students
            if not self.student_combo.get():
                self.student_combo.current(0)

        except Exception as e:
            print(f"Load Students Error: {e}")

    # ================= LOAD DATA =================
    def load_data(self):
        # FIX: guard against None current_user
        if not self.controller.current_user:
            return
        try:
            start       = self.start_date.get()
            end         = self.end_date.get()
            student     = self.student_combo.get()
            employee_id = self.controller.current_user.get("employee_id")
            if not employee_id:
                return

            with db_connect() as conn:
                with conn.cursor() as cur:
                    query = """
                        SELECT h.student_name, h.student_id, h.time_out,
                               h.fetcher_name, h.location
                        FROM history_log h
                        JOIN classroom c
                            ON h.student_id = c.student_id
                            AND c.employee_id = h.employee_id
                        WHERE h.employee_id = %s
                        AND DATE(h.time_out) BETWEEN %s AND %s
                    """
                    params = [employee_id, start, end]

                    if student and student != "ALL":
                        query += " AND h.student_name = %s"
                        params.append(student)

                    query += " ORDER BY h.time_out DESC"
                    cur.execute(query, tuple(params))
                    self.all_rows = cur.fetchall()

            self.update_table(self.all_rows)
            self._update_stats()

        except Exception as e:
            print(f"Load Data Error: {e}")

    # ================= UPDATE TABLE =================
    def update_table(self, rows):
        self.tree.delete(*self.tree.get_children())
        for i, row in enumerate(rows):
            tag = "odd" if i % 2 == 0 else "even"
            self.tree.insert("", "end", values=row, tags=(tag,))

    # ================= STATS =================
    def _update_stats(self):
        total   = len(self.all_rows)
        unique  = len(set(row[0] for row in self.all_rows)) if self.all_rows else 0
        self.total_label.config(text=f"Total Records: {total}")
        self.student_count.config(text=f"Unique Students: {unique}")
        self.status_label.config(text=f"Last refreshed — {total} record(s) loaded")

    # ================= LIVE SEARCH =================
    def live_search(self, *args):
        q = self.search_var.get().lower()
        filtered = [row for row in self.all_rows if q in str(row).lower()]
        self.update_table(filtered)

    # ================= EXPORT EXCEL =================
    def export_excel(self):
        if not self.all_rows:
            messagebox.showwarning("No Data", "Load data first before exporting.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save Student Logs"
        )
        if not file_path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Student Logs"

        headers = ["Student Name", "Student ID", "Time Out", "Fetcher", "Location"]
        header_fill = PatternFill("solid", fgColor="0047AB")
        header_font = Font(bold=True, color="FFFFFF", name="Segoe UI")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = 20

        for row in self.all_rows:
            ws.append(list(row))

        wb.save(file_path)
        messagebox.showinfo("Success", f"Exported {len(self.all_rows)} records successfully!")

    # ================= CHART =================
    def show_chart(self):
        if not self.all_rows:
            messagebox.showwarning("No Data", "Load data first before showing chart.")
            return

        names = [row[0] for row in self.all_rows]
        count = Counter(names)

        fig, ax = plt.subplots(figsize=(10, 5))
        bars = ax.bar(count.keys(), count.values(), color="#0047AB", edgecolor="white", linewidth=0.8)

        for bar in bars:
            ax.text(bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + 0.1,
                    str(int(bar.get_height())),
                    ha="center", va="bottom", fontsize=9, color="#374151")

        ax.set_title("Student Fetch Frequency", fontsize=14, fontweight="bold", color="#1F2937", pad=15)
        ax.set_xlabel("Student Name", fontsize=10, color="#374151")
        ax.set_ylabel("Times Fetched", fontsize=10, color="#374151")
        ax.set_facecolor("#F8FAFC")
        fig.patch.set_facecolor("#F0F4F8")
        plt.xticks(rotation=40, ha="right", fontsize=9)
        plt.tight_layout()
        plt.show()